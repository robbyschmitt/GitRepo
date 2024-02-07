# Takes the AWS (DV-Anwendungen und TIA) File from Data copies it to Dataoutput and sets it up for conversion.
# After the Script runs take the output and convert it to xml 

import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET

# First we import and configure the AWS File
def process_AWSexcel_file(input_AWSfile, output_AWSfile):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(input_AWSfile)

    for col_index, col_value in enumerate(df.columns):
    # Check if the cell is part of a merged range
      if df.iloc[0, col_index] is not None and isinstance(df.iloc[0, col_index], str):
        # Unmerge the cell
        df.iloc[0, col_index] = None


    # Delete the first two rows
    df = df.iloc[1:]

    # Delete specified columns (first, third, sixth, seventh, and all the rest)
    columns_to_drop = [2,]  + list(range(4, 45))
    df = df.drop(df.columns[columns_to_drop], axis=1)

    # Add a new column with cells containing the  "ApplicationComponent" as the second column
    df.insert(1, 'NewColumn', 'ApplicationComponent')

    # Save the modified DataFrame to a new Excel file
    df.to_excel(output_AWSfile, index=False)

# Then we import and configure the TIA File
def process_TIAexcel_file(input_TIAfile, output_TIAfile):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(input_TIAfile)

    for col_index, col_value in enumerate(df.columns):
    # Check if the cell is part of a merged range
      if df.iloc[0, col_index] is not None and isinstance(df.iloc[0, col_index], str):
        # Unmerge the cell
        df.iloc[0, col_index] = None

    
    # Delete the first two rows
    df = df.iloc[1:]

    # Delete specified columns 
    columns_to_drop = [2,]  + list(range(4, 17))
    df = df.drop(df.columns[columns_to_drop], axis=1)

    # Add a new column with cells containing the  "ApplicationComponent" as the second column
    df.insert(1, 'NewColumn', 'ApplicationComponent')

    # Save the modified DataFrame to a new Excel file
    df.to_excel(output_TIAfile, index=False)


def open_excel_sheet(output_AWSfile_name):
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(output_AWSfile_name)
    
    # Assuming you want to work with the first sheet (index 0)
    sheet = workbook.worksheets[0]
    
    return workbook, sheet

def create_xml_from_excel(sheet):
    root = ET.Element("elements")

    # Iterate through rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        element_identifier, xsi_type, name_de, documentation = row

        element = ET.SubElement(root, f"element", attrib={'identifier': element_identifier, 'xsi:type': xsi_type})

        name = ET.SubElement(element, f"name", attrib={'xml:lang': 'en'})
        name.text = name_de

        doc = ET.SubElement(element, f"documentation")
        doc.text = documentation


    return root

def save_xml(xml_root, xml_file_path1):
    tree = ET.ElementTree(xml_root)
    tree.write(xml_file_path1, encoding="utf-8", xml_declaration=True)
    print(f"XML file saved to {xml_file_path1} successfully.")

def convert1(output_AWSfile_name, xml_file_path1):
    try:
        workbook, sheet = open_excel_sheet(output_AWSfile_name)
        
        # Create XML from the Excel sheet
        xml_root = create_xml_from_excel(sheet)
        
        # Save XML to a file
        save_xml(xml_root, xml_file_path1)
    except Exception as e:
        print(f"Error: {e}")
    finally:
        # Close the workbook
        if 'workbook' in locals():
            workbook.close()


#########

def open_excel_sheet(output_TIAfile_name):
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(output_TIAfile_name)
    
    # Assuming you want to work with the first sheet (index 0)
    sheet = workbook.worksheets[0]
    
    return workbook, sheet

def create_xml_from_excel(sheet):
    root = ET.Element("elements")

    # Iterate through rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        element_identifier, xsi_type, name_de, documentation = row

        element = ET.SubElement(root, f"element", attrib={'identifier': element_identifier, 'xsi:type': xsi_type})

        name = ET.SubElement(element, f"name", attrib={'xml:lang': 'en'})
        name.text = name_de

        doc = ET.SubElement(element, f"documentation")
        doc.text = documentation


    return root

def save_xml(xml_root, xml_file_path2):
    tree = ET.ElementTree(xml_root)
    tree.write(xml_file_path2, encoding="utf-8", xml_declaration=True)
    print(f"XML file saved to {xml_file_path2} successfully.")

def convert2(output_TIAfile_name, xml_file_path2):
    try:
        workbook, sheet = open_excel_sheet(output_TIAfile_name)
        
        # Create XML from the Excel sheet
        xml_root = create_xml_from_excel(sheet)
        
        # Save XML to a file
        save_xml(xml_root, xml_file_path2)
    except Exception as e:
        print(f"Error: {e}")
    finally:
        # Close the workbook
        if 'workbook' in locals():
            workbook.close()



if __name__ == "__main__":
    # setup the file names
    input_AWSfile_name = 'u:\Temp\Programming\HLBDM\Data\AWSDB-AWS.xlsx'
    output_AWSfile_name = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-AWS-Elem-S1.xlsx'
    input_TIAfile_name = 'u:\Temp\Programming\HLBDM\Data\AWSDB-TIAs.xlsx'
    output_TIAfile_name = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-TIA-Elem-S1.xlsx'
    output_AWSTIA_name = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-AWSTIA.xlsx'
    file_path1 = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-AWS-Elem-S1.xlsx'
    xml_file_path1 = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-AWS.xml'
    xml_file_path2 = 'u:\Temp\Programming\HLBDM\Dataoutput\Imp-TIA.xml'

    process_AWSexcel_file(input_AWSfile_name, output_AWSfile_name)
    print(f"Excel AWS ile processed and saved as {output_AWSfile_name}")
    process_TIAexcel_file(input_TIAfile_name, output_TIAfile_name)
    print(f"Excel AWS ile processed and saved as {output_AWSfile_name}")
#    append_excel_rows(output_AWSfile_name, output_TIAfile_name, output_AWSTIA_name)
    print(f"Excel AWS ile processed and saved as {output_AWSfile_name}")
    convert1(output_AWSfile_name, xml_file_path1)
    convert2(output_TIAfile_name, xml_file_path2)

    