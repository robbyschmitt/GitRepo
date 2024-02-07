import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def compare_and_highlight(input_file1, input_file2, output_file, column_index1=0, column_index2=2, prefix_to_ignore="IS"):
    # Read Excel files into pandas dataframes
    df1 = pd.read_excel(input_file1)
    df2 = pd.read_excel(input_file2)

    # Extract columns based on specified indices from both dataframes
    df1_column = df1.columns[column_index1]
    df2_column = df2.columns[column_index2]

    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the new sheet
    ws.append(df1.columns.tolist())

    # Iterate through rows of the first dataframe
    for index, row in df1.iterrows():
        # Strip the prefix before comparison
        value_to_compare = str(row[df1_column]).lstrip(prefix_to_ignore)
        
        # Check if the value in the first column is present in the second dataframe's column
        if value_to_compare not in df2[df2_column].astype(str).str.lstrip(prefix_to_ignore).tolist():
            # If not present, highlight the row in red
            for col_num, value in enumerate(row):
                ws.cell(row=index + 2, column=col_num + 1, value=value).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            # If present, write the row without highlighting
            ws.append(row.tolist())
                        

    # Save the modified Excel file
    wb.save(output_file)

# Example usage
SourcePath = "u:\Temp\Programming\HLBDM\Dataoutput\Server-SW.xlsx"
TargetPath = "u:\Temp\Programming\HLBDM\Dataoutput\Server-Srv.xlsx"
output_file_path = "u:\Temp\Programming\HLBDM\Dataoutput\Diff.xlsx"
column_index1 = 0  # Index of the  column in the first sheet
column_index2 = 0  # Index of the  column in the second sheet
prefix_to_ignore = ""

compare_and_highlight(SourcePath, TargetPath, output_file_path, column_index1, column_index2, prefix_to_ignore)
print(f"Saved resilts to {output_file_path}")